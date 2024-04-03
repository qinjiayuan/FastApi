#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/4/1 17:18
# @Author  : qinjiayuan
# @File    : xmind_convert_excel.py


import datetime
import os
import xmindparser
from openpyxl import Workbook

'''
导入xmind文件
'''


class Xmind_to_excel():
    def __init__(self,xmind_path : str  , name : str):
        self.xmind_path : str = xmind_path   #xmind的文件路径
        self.name : str = name               #设计人员的名字

#递归获取xmind上的测试用例
    def get_testcase(self,data: dict, titles: list = None, prefix: str = "") -> list:
        if isinstance(data,dict):
            if titles is None:
                titles = []
            if 'title' in data:
                prefix += data["title"] + "-"
            if 'topics' in data:
                for topic in data['topics']:
                    self.get_testcase(topic, titles, prefix)
                prefix = prefix[:-1]  # Remove the last '-'
            else:
                titles.append(prefix[:-1])  # Remove the last '-'
            return titles

#将xmind转换成py对象
    def get_xmind_dict(self ) -> list:
        xmind_data = xmindparser.xmind_to_dict((self.xmind_path))
        # print(xmind_data[0]['topic'])
        result = self.get_testcase(xmind_data[0]['topic'])
        return result

#组装测试用例
    def make_testcase(self) -> list :
        test_dict = self.get_xmind_dict()
        test_case = []

#分离每个用例，并且存放到对应的列中
        for i in test_dict :
            demo_list = []
            name = []
            step = []
            expection = []
            step_name = []
            data = i.split('step')
            name.append(data[0])
            data = data[1].split('expection')
            expection.append(data[1])
            data = data[0].split('：')
            step_name.append(data[0])
            step.append(data[1])

            demo_list.append(name[0][:-1])
            demo_list.append(step_name[0])
            demo_list.append(step[0][:-1])
            demo_list.append(expection[0][:-1])
            test_case.append(demo_list)
        print(f'测试用例生成：{test_case}')
        return test_case


    def write_to_excel(self) -> None:
        test_case = self.make_testcase()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sheet1'

        header = ['案例目录', '案例编号', '测试概述', '前置条件', '测试数据', '操作步骤名称', '操作步骤',
                  '预期结果', '备注', '设计人员', '案例类型', '优先级', '设计日期', '自动化状态', '标签类型',
                  '一站式用例编号（非必填）']
        need_merge_row = ['A', 'B', 'C', 'D', 'E', 'I', 'J', 'K', 'L', 'M', 'N']

        # 将标题行添加到工作表中
        for col_num, header_text in enumerate(header, 1):
            worksheet.cell(row=1, column=col_num, value=header_text)

        row_num = 2
        for case in test_case:
            worksheet.cell(row=row_num, column=3, value=case[0])        # 测试概述
            worksheet.cell(row=row_num,column=6,value=case[1])          #操作步骤名称
            worksheet.cell(row=row_num, column=7, value=case[2])        # 操作步骤
            worksheet.cell(row=row_num, column=8, value=case[3])        # 预期结果
            worksheet.cell(row=row_num, column=10, value=self.name)     # 设计人员
            worksheet.cell(row=row_num, column=11, value='正案例')        # 案例类型
            worksheet.cell(row=row_num, column=12, value='中')           # 优先级
            worksheet.cell(row=row_num, column=13, value=f'{datetime.date.today()}')  # 设计日期
            row_num += 1

#合并测试概述相同的单元格
        value_same = 2
        while value_same < worksheet.max_row:  # value_same 从excle 的第二行开始取并且 < excel 的 最大行数
            index =  value_same + 1
            while index <= worksheet.max_row + 1:

                # print(f'第{value_same}行：{worksheet.cell(row=value_same,column=3).value}')
                # print(f'第{index}行：{worksheet.cell(row=index,column=3).value}')
                if worksheet.cell(row=value_same,column=3).value == worksheet.cell(row=index,column=3).value: #判断value_same行的测试概述和index的测试概述是否相同
                    index += 1

                elif index - value_same > 1 or index >= worksheet.max_row :
                    for row_num in need_merge_row:
                        worksheet.merge_cells(f'{row_num}{value_same}:{row_num}{index - 1}')
                    value_same = index
                    break

                elif index - value_same == 1: #若前后两行的测试概述不同，直接跳过第value_same行
                    value_same = index
                    break

#删除操作过的最后一行，以便导入成功
        worksheet.delete_rows(value_same)

#设置用例附件的格式
        today_original = datetime.datetime.today()
        today = datetime.datetime.strftime(today_original,'%Y_%m_%d')
        workbook.save(f'{self.name}_testcase_{today}.xlsx')
        print('测试用例成功生成')


if __name__ == '__main__':
    '''
     xmind当中 如果是步骤必须加上'step', 如“步骤1” 要写成“step步骤1”且步骤后面必须接中文冒号，“预期”必须写成“expection预期1”,等，每个连续用例的合并是根据测试概述来合并，如果测试概述不同，则不连续的则不会合并
    '''
    xmind_path =  r'C:\Users\EDY\Desktop\Xmind\2024\测试tab.xmind'  #附件路径
    name = 'wxqinjiayuan' #填入自己的名字
    demo = Xmind_to_excel(xmind_path,name)
    result = demo.write_to_excel()
